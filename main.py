# This is a sample Python script.

# Press <no shortcut> to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.

from openpyxl import Workbook
from excel_object import ClusterEdit
from const import filled_excel_path, stable_db_path
from docx import Document


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press <no shortcut> to toggle the breakpoint.


def save_excel(data_list):
    workbook = Workbook()
    mysheet = workbook.active
    mysheet.title = 'sheet2'
    mysheet['A1'] = '哈哈'
    workbook.save('123.xlsx')


def read_excel(startline, endline, path):
    from openpyxl import load_workbook

    wb = load_workbook(path)
    sheet = wb.active
    print(sheet)
    cluster_data = []
    for line in range(startline, endline):
        if not sheet['A' + str(line)].value:
            break
        edit_data = ClusterEdit(sheet['A' + str(line)].value, sheet['B' + str(line)].value,
                                sheet['C' + str(line)].value, sheet['D' + str(line)].value)
        cluster_data.append(edit_data)
    return cluster_data


def print_list_object(object_list):
    for obj in object_list:
        print(obj.__dict__)


def write_anquanjian_word(anquanjian_list):
    document = Document()
    table = document.add_table(rows=len(anquanjian_list) + 2, cols=5, style='Table Grid')
    table.cell(0, 0).text = '产品描述\n6.安全件一览表'
    table.cell(0, 0).merge(table.cell(0, 4))
    
    table.cell(1, 1).text = '元/部件名称'
    table.cell(1, 2).text = '元件材料名称'
    table.cell(1, 3).text = '型号规格/牌号'
    table.cell(1, 4).text = '制造商（生产厂）'
    
    a_count = 1
    a_start_line = 2
    b_start_line = 2
    c_start_line = 2

    table.cell(2, 0).text = str(a_count)
    table.cell(2, 1).text = anquanjian_list[0].a
    table.cell(2, 2).text = anquanjian_list[0].b
    table.cell(2, 3).text = anquanjian_list[0].gui_ge
    table.cell(2, 4).text = anquanjian_list[0].factory

    print(len(anquanjian_list))
    for i in range(3, len(anquanjian_list) + 2):
        print(anquanjian_list[i-2].__dict__)
        
        table.cell(i, 4).text = anquanjian_list[i - 2].factory

        now_gui_ge = anquanjian_list[i - 2].gui_ge
        now_b = anquanjian_list[i - 2].b
        now_a = anquanjian_list[i - 2].a
        if now_gui_ge != anquanjian_list[c_start_line - 2].gui_ge:
            if i - 1 != c_start_line:
                table.cell(c_start_line, 3).merge(table.cell(i - 1, 3))
            c_start_line = i
            table.cell(i, 3).text = anquanjian_list[i - 2].gui_ge
        
        if now_b != anquanjian_list[b_start_line - 2].b:
            if i - 1 != b_start_line:
                table.cell(b_start_line, 2).merge(table.cell(i - 1, 2))
            b_start_line = i
            table.cell(i, 2).text = anquanjian_list[i - 2].b
        
        if now_a != anquanjian_list[a_start_line - 2].a:
            if i - 1 != a_start_line:
                table.cell(a_start_line, 1).merge(table.cell(i - 1, 1))
                table.cell(a_start_line, 0).merge(table.cell(i - 1, 0))
                a_count += 1
                
            a_start_line = i
            table.cell(i, 1).text = anquanjian_list[i - 2].a
            table.cell(i, 0).text = str(a_count)

    last_line = len(anquanjian_list) + 1
    if anquanjian_list[last_line - 3].gui_ge == anquanjian_list[c_start_line - 2].gui_ge:
        table.cell(c_start_line, 3).merge(table.cell(last_line, 3))
    if anquanjian_list[last_line - 3].b == anquanjian_list[b_start_line - 2].b:
        table.cell(b_start_line, 2).merge(table.cell(last_line, 2))
    if anquanjian_list[last_line - 3].a == anquanjian_list[a_start_line - 2].a:
        table.cell(a_start_line, 1).merge(table.cell(last_line, 1))
        table.cell(a_start_line, 0).merge(table.cell(last_line, 0))

    document.save('123.docx')


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print_hi('PyCharm')

cluster_data = read_excel(3, 27, filled_excel_path)
stable_data = read_excel(2, 200, stable_db_path)

merge_result = []

# 创建用户新增数据的字典 key:a列加b列 value:数据list
if cluster_data:
    cluster_data_dict = {}
    for one_data in cluster_data:
        data_list = cluster_data_dict.get(one_data.a + one_data.b)
        if data_list:
            data_list.append(one_data)
        else:
            cluster_data_dict[one_data.a + one_data.b] = [one_data]

    wait_to_insert = False
    key = ''

    each_stable = next(stable_data.__iter__())
    for each_stable in stable_data:
        dict_cluster_data = cluster_data_dict.get(each_stable.a + each_stable.b)

        # 当等待插入，且key改变时，插入数据
        if wait_to_insert & (each_stable.a + each_stable.b != key):
            merge_result.extend(cluster_data_dict.get(key))
            wait_to_insert = False
            key = ''
        elif dict_cluster_data:
            wait_to_insert = True
            key = each_stable.a + each_stable.b
        merge_result.append(each_stable)
else:
    merge_result = stable_data

print_list_object(merge_result)
write_anquanjian_word(merge_result)